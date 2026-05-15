from functools import wraps
from io import BytesIO
import re
import secrets

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.conf import settings
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from .forms import (
    AccountSettingsForm,
    AdoptionApplicationAdminForm,
    AdoptionApplicationForm,
    FeedbackForm,
    NewsAdminForm,
    OwnerQuestionnaireForm,
    PetAdminForm,
    PetImportForm,
    RegisterForm,
)
from .models import AdoptionApplication, Feedback, News, OwnerQuestionnaire, Pet, UserProfile


def ensure_user_profile(user):
    if not user.is_authenticated:
        return None
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile


def get_effective_role(user):
    if not user.is_authenticated:
        return None
    if user.is_superuser or user.is_staff:
        return 'admin'
    profile = ensure_user_profile(user)
    return profile.role if profile else 'user'


def admin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if get_effective_role(request.user) != 'admin':
            messages.error(request, 'Эта страница доступна только администратору.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)

    return wrapper


def sync_user_related_records(user):
    if not user.is_authenticated:
        return

    ensure_user_profile(user)

    if user.email:
        AdoptionApplication.objects.filter(user__isnull=True, email=user.email).update(user=user)

    full_name = user.get_full_name().strip()
    if full_name:
        OwnerQuestionnaire.objects.filter(user__isnull=True, full_name=full_name).update(user=user)


def build_temporary_username(email):
    local_part = email.split('@', 1)[0].strip().lower()
    base = re.sub(r'[^0-9a-zA-Z_]+', '', local_part) or 'user'
    username = base[:20]
    counter = 1

    while User.objects.filter(username=username).exists():
        suffix = str(counter)
        username = f'{base[: max(1, 20 - len(suffix))]}{suffix}'
        counter += 1

    return username


def create_temporary_account(email, full_name=''):
    existing_user = User.objects.filter(email__iexact=email).first()
    if existing_user:
        return existing_user, None, False, True

    password = secrets.token_urlsafe(8)
    username = build_temporary_username(email)
    first_name = full_name.strip().split()[0] if full_name.strip() else ''

    user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
    )
    ensure_user_profile(user)
    sync_user_related_records(user)
    return user, password, True, False


def temporary_account_email_is_configured():
    backend = getattr(settings, 'EMAIL_BACKEND', '')
    non_delivery_backends = {
        'django.core.mail.backends.console.EmailBackend',
        'django.core.mail.backends.filebased.EmailBackend',
        'django.core.mail.backends.locmem.EmailBackend',
        'django.core.mail.backends.dummy.EmailBackend',
    }

    if backend in non_delivery_backends:
        return False

    if backend == 'django.core.mail.backends.smtp.EmailBackend':
        return bool(getattr(settings, 'EMAIL_HOST', '') and getattr(settings, 'DEFAULT_FROM_EMAIL', ''))

    return bool(getattr(settings, 'DEFAULT_FROM_EMAIL', ''))


def send_temporary_account_email(user, email, password):
    if not temporary_account_email_is_configured():
        return False

    subject = 'Временный аккаунт на сайте приюта "Лапки"'
    message = (
        f'Здравствуйте!\n\n'
        f'Для вас создан временный аккаунт на сайте приюта "Лапки".\n\n'
        f'Логин: {user.username}\n'
        f'Пароль: {password}\n\n'
        f'После входа вы сможете изменить данные в личном кабинете и посмотреть свои анкеты и заявки.\n\n'
        f'Если вы не запрашивали создание аккаунта, просто проигнорируйте это письмо.'
    )

    try:
        return send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        ) > 0
    except Exception:
        return False


QUESTIONNAIRE_COMPARE_FIELDS = [
    'full_name',
    'age',
    'city',
    'housing_type',
    'has_children',
    'has_other_pets',
    'experience_years',
    'activity_level',
    'time_at_home',
    'pet_preference',
    'pet_age_preference',
    'pet_gender_preference',
    'adoption_goal',
    'ready_for_medical_care',
    'additional_info',
]


def get_questionnaire_payload(source):
    return {field: getattr(source, field) for field in QUESTIONNAIRE_COMPARE_FIELDS}


def questionnaires_match(source, cleaned_data):
    for field in QUESTIONNAIRE_COMPARE_FIELDS:
        if getattr(source, field) != cleaned_data.get(field):
            return False
    return True


def build_questionnaire_initial(questionnaire):
    return {
        'full_name': questionnaire.full_name,
        'age': questionnaire.age,
        'city': questionnaire.city,
        'housing_type': questionnaire.housing_type,
        'has_children': questionnaire.has_children,
        'has_other_pets': questionnaire.has_other_pets,
        'has_pet_experience': questionnaire.experience_years > 0,
        'experience_years': questionnaire.experience_years,
        'activity_level': questionnaire.activity_level,
        'time_at_home': questionnaire.time_at_home,
        'pet_preference': questionnaire.pet_preference,
        'pet_age_preference': questionnaire.pet_age_preference,
        'pet_gender_preference': questionnaire.pet_gender_preference,
        'adoption_goal': questionnaire.adoption_goal,
        'ready_for_medical_care': questionnaire.ready_for_medical_care,
        'additional_info': questionnaire.additional_info,
    }


def build_matches_for_questionnaire(questionnaire):
    questionnaire_prefill = {
        'full_name': questionnaire.full_name,
        'age': questionnaire.age,
        'city': questionnaire.city,
        'housing_type': questionnaire.housing_type,
        'has_children': questionnaire.has_children,
        'has_other_pets': questionnaire.has_other_pets,
        'experience_years': questionnaire.experience_years,
    }

    pets = Pet.objects.filter(adopted=False)
    all_matches = []

    for pet in pets:
        score = calculate_match_score(questionnaire, pet)
        all_matches.append(
            {
                'pet': pet,
                'score': score,
                'label': get_match_label(score),
                'reasons': get_match_reasons(questionnaire, pet),
                'warnings': get_match_warnings(questionnaire, pet),
            }
        )

    all_matches.sort(key=lambda item: item['score'], reverse=True)
    matched_pets = [item for item in all_matches if item['score'] >= 45]
    used_fallback = False

    if not matched_pets:
        matched_pets = all_matches
        used_fallback = True

    return {
        'questionnaire': questionnaire,
        'matched_pets': matched_pets,
        'used_fallback': used_fallback,
        'questionnaire_prefill': questionnaire_prefill,
    }


def calculate_match_score(questionnaire, pet):
    score = 0

    if questionnaire.pet_preference == 'any':
        score += 10
    elif questionnaire.pet_preference == pet.type:
        score += 20
    else:
        score -= 15

    if questionnaire.pet_age_preference == 'any':
        score += 5
    elif questionnaire.pet_age_preference == pet.age_group:
        score += 10
    else:
        score -= 5

    if questionnaire.pet_gender_preference == 'any':
        score += 5
    elif questionnaire.pet_gender_preference == pet.gender:
        score += 8
    else:
        score -= 3

    if questionnaire.has_children:
        score += 15 if pet.can_live_with_children else -25
    else:
        score += 5

    if questionnaire.has_other_pets:
        score += 15 if pet.can_live_with_other_pets else -25
    else:
        score += 5

    if questionnaire.activity_level == pet.activity_level:
        score += 15
    elif (
        questionnaire.activity_level == 'medium' and pet.activity_level in ['low', 'high']
    ) or (pet.activity_level == 'medium' and questionnaire.activity_level in ['low', 'high']):
        score += 7
    else:
        score -= 5

    if questionnaire.housing_type == 'apartment':
        score += 12 if pet.suitable_for_apartment else -12
    elif questionnaire.housing_type == 'rented':
        score += -5 if pet.requires_experience or pet.health_status == 'care_needed' else 6
    else:
        score += 8

    if pet.requires_experience:
        score += 12 if questionnaire.experience_years >= 1 else -15
    else:
        score += 8

    if questionnaire.time_at_home == 'always':
        score += 10
    elif questionnaire.time_at_home == 'often':
        score += 7
    elif questionnaire.time_at_home == 'rarely':
        score += 5 if pet.can_stay_alone else -10

    if pet.health_status == 'care_needed':
        score += 15 if questionnaire.ready_for_medical_care else -20
    else:
        score += 8

    if questionnaire.adoption_goal == 'any':
        score += 5
    elif questionnaire.adoption_goal == 'family':
        score += 10 if pet.can_live_with_children and pet.temperament in ['friendly', 'calm'] else -5
    elif questionnaire.adoption_goal == 'active_walks':
        score += 12 if pet.type == 'dog' and pet.activity_level == 'high' else -5
    elif questionnaire.adoption_goal == 'companion':
        score += 10 if pet.temperament in ['friendly', 'calm'] else 3

    return max(0, min(score, 100))


def get_match_reasons(questionnaire, pet):
    reasons = []

    if questionnaire.pet_preference == pet.type:
        reasons.append('Соответствует предпочтению по виду')
    if questionnaire.pet_age_preference == pet.age_group:
        reasons.append('Подходит по возрасту')
    if questionnaire.pet_gender_preference == pet.gender:
        reasons.append('Подходит по полу')
    if questionnaire.has_children and pet.can_live_with_children:
        reasons.append('Подходит для семьи с детьми')
    if questionnaire.has_other_pets and pet.can_live_with_other_pets:
        reasons.append('Можно с другими животными')
    if questionnaire.activity_level == pet.activity_level:
        reasons.append('Совпадает уровень активности')
    if questionnaire.housing_type == 'apartment' and pet.suitable_for_apartment:
        reasons.append('Подходит для квартиры')
    if not pet.requires_experience:
        reasons.append('Подойдёт без опыта')
    if pet.health_status == 'healthy':
        reasons.append('Не требует особого ухода')
    elif pet.health_status == 'care_needed' and questionnaire.ready_for_medical_care:
        reasons.append('Вы готовы к особенностям здоровья')

    if questionnaire.adoption_goal == 'family' and pet.can_live_with_children:
        reasons.append('Подходит для семьи')
    elif questionnaire.adoption_goal == 'active_walks' and pet.activity_level == 'high':
        reasons.append('Подходит для активного образа жизни')
    elif questionnaire.adoption_goal == 'companion' and pet.temperament in ['calm', 'friendly']:
        reasons.append('Хороший компаньон')

    return reasons[:5]


def get_match_warnings(questionnaire, pet):
    warnings = []

    if questionnaire.has_children and not pet.can_live_with_children:
        warnings.append('Не подходит для семьи с детьми')
    if questionnaire.has_other_pets and not pet.can_live_with_other_pets:
        warnings.append('Может не ужиться с другими животными')
    if questionnaire.housing_type == 'apartment' and not pet.suitable_for_apartment:
        warnings.append('Не лучший вариант для квартиры')
    if pet.requires_experience and questionnaire.experience_years < 1:
        warnings.append('Желателен опыт содержания животных')
    if pet.health_status == 'care_needed' and not questionnaire.ready_for_medical_care:
        warnings.append('Потребуется особый уход')
    if questionnaire.time_at_home == 'rarely' and not pet.can_stay_alone:
        warnings.append('Тяжело переносит долгое одиночество')

    return warnings[:4]


def get_match_label(score):
    if score >= 85:
        return 'Очень высокая совместимость'
    if score >= 70:
        return 'Хорошая совместимость'
    return 'Средняя совместимость'


def account_home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'shelter/account_home.html')


def register(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            sync_user_related_records(user)
            login(request, user)
            messages.success(request, 'Аккаунт создан. Добро пожаловать!')
            return redirect('dashboard')
    else:
        form = RegisterForm()

    return render(request, 'shelter/register.html', {'form': form})


@login_required
def dashboard(request):
    sync_user_related_records(request.user)
    role = get_effective_role(request.user)

    account_form = AccountSettingsForm(instance=request.user)

    if request.method == 'POST':
        account_form = AccountSettingsForm(request.POST, instance=request.user)
        if account_form.is_valid():
            user = account_form.save()
            if account_form.cleaned_data.get('password1'):
                update_session_auth_hash(request, user)
            messages.success(request, 'Данные аккаунта обновлены.')
            return redirect('dashboard')

    owner_applications = request.user.adoption_applications.select_related('pet').order_by('-created_at')
    owner_questionnaires = request.user.owner_questionnaires.order_by('-created_at')

    context = {
        'role': role,
        'account_form': account_form,
        'owner_applications': owner_applications[:10],
        'owner_questionnaires': owner_questionnaires[:10],
        'owner_applications_count': owner_applications.count(),
        'owner_questionnaires_count': owner_questionnaires.count(),
    }

    if role == 'admin':
        context.update(
            {
                'users_count': User.objects.count(),
                'pets_count': Pet.objects.count(),
                'feedback_count': Feedback.objects.count(),
                'all_applications_count': AdoptionApplication.objects.count(),
                'all_questionnaires_count': OwnerQuestionnaire.objects.count(),
                'news_count': News.objects.count(),
                'adopted_count': Pet.objects.filter(adopted=True).count(),
                'recent_applications': AdoptionApplication.objects.select_related('pet', 'user').order_by('-created_at')[:6],
                'recent_questionnaires': OwnerQuestionnaire.objects.select_related('user').order_by('-created_at')[:6],
                'recent_feedback': Feedback.objects.order_by('-submitted_at')[:6],
                'recent_pets': Pet.objects.order_by('-id')[:6],
                'recent_news': News.objects.order_by('-date')[:4],
            }
        )

    return render(request, 'shelter/dashboard.html', context)


def owner_questionnaire(request):
    latest_questionnaire = None
    if request.user.is_authenticated:
        latest_questionnaire = request.user.owner_questionnaires.order_by('-created_at').first()

    if (
        request.method == 'GET'
        and request.user.is_authenticated
        and latest_questionnaire
        and request.GET.get('reuse') == '1'
    ):
        return render(
            request,
            'shelter/match_results.html',
            build_matches_for_questionnaire(latest_questionnaire),
        )

    if request.method == 'POST':
        form = OwnerQuestionnaireForm(request.POST, user_is_authenticated=request.user.is_authenticated)
        if form.is_valid():
            questionnaire = None
            save_new_questionnaire = True
            normalized_experience = 1 if form.cleaned_data.get('has_pet_experience') else 0

            if request.user.is_authenticated and latest_questionnaire and questionnaires_match(
                latest_questionnaire,
                {**form.cleaned_data, 'experience_years': normalized_experience},
            ):
                questionnaire = latest_questionnaire
                save_new_questionnaire = False

            if questionnaire is None:
                questionnaire = form.save(commit=False)
                questionnaire.experience_years = normalized_experience
            else:
                questionnaire.experience_years = normalized_experience

            if request.user.is_authenticated:
                questionnaire.user = request.user
            else:
                contact_email = form.cleaned_data.get('contact_email')
                if contact_email:
                    temp_user, temp_password, created, already_exists = create_temporary_account(
                        contact_email,
                        form.cleaned_data.get('full_name', ''),
                    )
                    if already_exists:
                        messages.warning(
                            request,
                            'Аккаунт с такой почтой уже существует. Анкета сохранена, но чтобы видеть её в кабинете, войдите под своим логином.',
                        )
                    else:
                        questionnaire.user = temp_user
                        login(request, temp_user)
                        email_sent = send_temporary_account_email(
                            temp_user,
                            contact_email,
                            temp_password,
                        )
                        if email_sent:
                            messages.success(
                                request,
                                'Мы создали временный аккаунт и отправили логин с паролем на указанную почту.',
                            )
                        else:
                            messages.success(
                                request,
                                f'Мы создали временный аккаунт. Логин: {temp_user.username}. Пароль: {temp_password}. Их можно изменить позже в кабинете.',
                            )
                else:
                    messages.info(
                        request,
                        'Анкета сохранена без аккаунта. Если хотите видеть её в кабинете, в следующий раз войдите в аккаунт или укажите email.',
                    )

            if save_new_questionnaire:
                questionnaire.save()
                if request.user.is_authenticated and latest_questionnaire:
                    messages.success(request, 'Изменённая анкета сохранена.')
            else:
                messages.info(request, 'Эта анкета уже сохранена. Показываем актуальный подбор.')

            return render(
                request,
                'shelter/match_results.html',
                build_matches_for_questionnaire(questionnaire),
            )
    else:
        initial = {}
        if request.user.is_authenticated:
            initial = {
                'full_name': request.user.get_full_name(),
                'contact_email': request.user.email,
            }
            if latest_questionnaire and request.GET.get('edit') == '1':
                initial.update(build_questionnaire_initial(latest_questionnaire))
        form = OwnerQuestionnaireForm(initial=initial, user_is_authenticated=request.user.is_authenticated)

    return render(
        request,
        'shelter/owner_questionnaire.html',
        {
            'form': form,
            'show_saved_questionnaire_prompt': bool(
                request.user.is_authenticated
                and latest_questionnaire
                and request.method == 'GET'
                and request.GET.get('edit') != '1'
                and request.GET.get('reuse') != '1'
            ),
            'latest_questionnaire': latest_questionnaire,
        },
    )


def adoption_application(request, pet_id):
    pet = get_object_or_404(Pet, id=pet_id)

    if request.method == 'POST':
        form = AdoptionApplicationForm(request.POST)
        if form.is_valid():
            application = form.save(commit=False)
            application.pet = pet
            feedback_message = 'Заявка успешно отправлена. Мы свяжемся с будущим хозяином по указанной электронной почте.'
            if request.user.is_authenticated:
                application.user = request.user
            else:
                temp_user, temp_password, created, already_exists = create_temporary_account(
                    form.cleaned_data['email'],
                    form.cleaned_data.get('full_name', ''),
                )
                if already_exists:
                    feedback_message = 'Заявка отправлена. Чтобы видеть её в кабинете, войдите под своим логином.'
                else:
                    application.user = temp_user
                    login(request, temp_user)
                    email_sent = send_temporary_account_email(
                        temp_user,
                        form.cleaned_data['email'],
                        temp_password,
                    )
                    if email_sent:
                        feedback_message = 'Заявка отправлена. Мы создали временный аккаунт и отправили логин с паролем на указанную почту.'
                    else:
                        feedback_message = f'Заявка отправлена. Мы создали временный аккаунт. Логин: {temp_user.username}. Пароль: {temp_password}. Их можно изменить позже в кабинете.'
            application.save()
            messages.success(request, feedback_message)
            return redirect(f"{reverse('pets')}?application_sent=1")
    else:
        initial = {}
        if request.user.is_authenticated:
            latest_questionnaire = request.user.owner_questionnaires.order_by('-created_at').first()
            latest_application = request.user.adoption_applications.order_by('-created_at').first()
            initial = {
                'full_name': request.user.get_full_name(),
                'email': request.user.email,
            }
            if latest_application:
                initial.update(
                    {
                        'phone': latest_application.phone,
                    }
                )
            if latest_questionnaire:
                housing_type_map = {
                    'apartment': 'Квартира',
                    'house': 'Частный дом',
                    'rented': 'Съёмное жильё',
                }
                initial.update(
                    {
                        'age': latest_questionnaire.age,
                        'city': latest_questionnaire.city,
                        'housing_type': housing_type_map.get(latest_questionnaire.housing_type, ''),
                        'has_other_pets': latest_questionnaire.has_other_pets,
                        'has_children': latest_questionnaire.has_children,
                        'experience': (
                            'Есть опыт содержания животных.'
                            if latest_questionnaire.experience_years >= 1
                            else 'Опыта содержания животных раньше не было.'
                        ),
                    }
                )
            elif latest_application:
                initial.update(
                    {
                        'age': latest_application.age,
                        'city': latest_application.city,
                        'housing_type': latest_application.housing_type,
                        'has_other_pets': latest_application.has_other_pets,
                        'has_children': latest_application.has_children,
                        'experience': latest_application.experience,
                    }
                )
        form = AdoptionApplicationForm(initial=initial)

    return render(request, 'shelter/adoption_application.html', {'form': form, 'pet': pet})


def home(request):
    latest_news = News.objects.all()[:3]
    happy_stories = Pet.objects.filter(adopted=True)[:3]

    context = {
        'latest_news': latest_news,
        'happy_stories': happy_stories,
        'total_pets': Pet.objects.count(),
        'adopted_pets': Pet.objects.filter(adopted=True).count(),
        'available_pets': Pet.objects.filter(adopted=False).count(),
        'applications_count': AdoptionApplication.objects.count(),
        'questionnaires_count': OwnerQuestionnaire.objects.count(),
    }
    return render(request, 'shelter/home.html', context)


def news_list(request):
    news_items = News.objects.order_by('-date')
    return render(request, 'shelter/news_list.html', {'news_list': news_items})


def news_detail(request, pk):
    news_item = get_object_or_404(News, pk=pk)
    return render(request, 'shelter/news_detail.html', {'news_item': news_item})


def pets_view(request):
    pets = Pet.objects.all()
    return render(
        request,
        'shelter/pets.html',
        {
            'pets': pets,
            'cats_count': pets.filter(type='cat').count(),
            'dogs_count': pets.filter(type='dog').count(),
        },
    )


def about(request):
    return render(request, 'shelter/about.html')


def help_page(request):
    initial_data = {}
    message_text = request.GET.get('message')
    if message_text:
        initial_data['message'] = message_text

    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Сообщение отправлено!')
            return redirect('help')
    else:
        form = FeedbackForm(initial=initial_data)

    return render(request, 'shelter/help.html', {'form': form})


def contacts(request):
    initial_data = {}
    message_text = request.GET.get('message')
    if message_text:
        initial_data['message'] = message_text

    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Сообщение отправлено!')
            return redirect('contacts')
    else:
        form = FeedbackForm(initial=initial_data)

    return render(request, 'shelter/contacts.html', {'form': form})


@admin_required
def admin_user_list(request):
    users = User.objects.select_related('userprofile').order_by('-date_joined')
    return render(
        request,
        'shelter/admin_user_list.html',
        {
            'users': users,
            'page_title': 'Пользователи',
        },
    )


PET_EXCEL_COLUMNS = [
    {
        'key': 'id',
        'hint': 'Оставьте пустым для нового питомца. Если указать существующий id, карточка обновится.',
    },
    {
        'key': 'name',
        'hint': 'Имя питомца. Например: Марта.',
    },
    {
        'key': 'type',
        'hint': 'Вид питомца: Кошка или Собака. Подойдут и значения cat / dog.',
    },
    {
        'key': 'gender',
        'hint': 'Пол питомца: Мальчик или Девочка. Подойдут и male / female.',
    },
    {
        'key': 'age_months',
        'hint': 'Возраст в месяцах. Например: 14.',
    },
    {
        'key': 'age_group',
        'hint': 'Возрастная группа: Молодой или Взрослый. Подойдут и young / adult.',
    },
    {
        'key': 'description',
        'hint': 'Короткая история или описание питомца.',
    },
    {
        'key': 'sterilized',
        'hint': 'Стерилизован: Да или Нет. Подойдут и true / false, 1 / 0.',
    },
    {
        'key': 'vaccinated',
        'hint': 'Привит: Да или Нет. Подойдут и true / false, 1 / 0.',
    },
    {
        'key': 'adopted',
        'hint': 'Пристроен ли питомец: Да или Нет.',
    },
    {
        'key': 'can_live_with_children',
        'hint': 'Можно ли в семью с детьми: Да или Нет.',
    },
    {
        'key': 'can_live_with_other_pets',
        'hint': 'Можно ли с другими животными: Да или Нет.',
    },
    {
        'key': 'activity_level',
        'hint': 'Уровень активности: Низкая, Средняя или Высокая. Подойдут и low / medium / high.',
    },
    {
        'key': 'suitable_for_apartment',
        'hint': 'Подходит ли для квартиры: Да или Нет.',
    },
    {
        'key': 'requires_experience',
        'hint': 'Нужен ли опытный хозяин: Да или Нет.',
    },
    {
        'key': 'temperament',
        'hint': 'Характер: Спокойный, Активный, Дружелюбный или Осторожный.',
    },
    {
        'key': 'health_status',
        'hint': 'Состояние здоровья: Здоров или Нужен особый уход.',
    },
    {
        'key': 'can_stay_alone',
        'hint': 'Может ли оставаться один: Да или Нет.',
    },
]

PET_EXCEL_HEADERS = [column['key'] for column in PET_EXCEL_COLUMNS]

PET_EXCEL_EXAMPLE = {
    'id': '',
    'name': 'Марта',
    'type': 'Кошка',
    'gender': 'Девочка',
    'age_months': 14,
    'age_group': 'Взрослый',
    'description': 'Спокойная кошка, любит людей и мягкие лежанки.',
    'sterilized': 'Да',
    'vaccinated': 'Да',
    'adopted': 'Нет',
    'can_live_with_children': 'Да',
    'can_live_with_other_pets': 'Да',
    'activity_level': 'Средняя',
    'suitable_for_apartment': 'Да',
    'requires_experience': 'Нет',
    'temperament': 'Дружелюбный',
    'health_status': 'Здоров',
    'can_stay_alone': 'Да',
}

PET_EXCEL_CHOICE_MAPPINGS = {
    'type': {
        'cat': 'cat',
        'кошка': 'cat',
        'dog': 'dog',
        'собака': 'dog',
    },
    'gender': {
        'male': 'male',
        'мальчик': 'male',
        'самец': 'male',
        'female': 'female',
        'девочка': 'female',
        'самка': 'female',
    },
    'age_group': {
        'young': 'young',
        'молодой': 'young',
        'adult': 'adult',
        'взрослый': 'adult',
    },
    'activity_level': {
        'low': 'low',
        'низкая': 'low',
        'medium': 'medium',
        'средняя': 'medium',
        'high': 'high',
        'высокая': 'high',
    },
    'temperament': {
        'calm': 'calm',
        'спокойный': 'calm',
        'active': 'active',
        'активный': 'active',
        'friendly': 'friendly',
        'дружелюбный': 'friendly',
        'careful': 'careful',
        'осторожный': 'careful',
    },
    'health_status': {
        'healthy': 'healthy',
        'здоров': 'healthy',
        'care_needed': 'care_needed',
        'нужен особый уход': 'care_needed',
        'особый уход': 'care_needed',
    },
}

PET_EXCEL_TRUE_VALUES = {'1', 'true', 'yes', 'да'}
PET_EXCEL_FALSE_VALUES = {'0', 'false', 'no', 'нет'}


def pet_excel_bool_label(value):
    return 'Да' if value else 'Нет'


def pet_excel_normalize_choice(value, mapping, default):
    raw_value = str(value).strip().lower()
    if not raw_value:
        return default
    return mapping.get(raw_value, default)


def pet_excel_bool_from_value(value, default=False):
    if value in (None, ''):
        return default
    if isinstance(value, bool):
        return value

    raw_value = str(value).strip().lower()
    if raw_value in PET_EXCEL_TRUE_VALUES:
        return True
    if raw_value in PET_EXCEL_FALSE_VALUES:
        return False
    return default


def build_pet_excel_context(form=None):
    form = form or PetImportForm()
    template_columns = [
        {
            'header': column['key'],
            'hint': column['hint'],
            'example': PET_EXCEL_EXAMPLE[column['key']],
        }
        for column in PET_EXCEL_COLUMNS
    ]

    return {
        'form': form,
        'page_title': 'Импорт и выгрузка Excel',
        'template_columns': template_columns,
        'field_hints': [
            ('type', 'Можно писать Кошка / Собака или cat / dog'),
            ('gender', 'Можно писать Мальчик / Девочка или male / female'),
            ('age_group', 'Можно писать Молодой / Взрослый или young / adult'),
            ('activity_level', 'Можно писать Низкая / Средняя / Высокая или low / medium / high'),
            ('temperament', 'Можно писать Спокойный, Активный, Дружелюбный или Осторожный'),
            ('health_status', 'Можно писать Здоров или Нужен особый уход'),
            ('поля Да/Нет', 'Поддерживаются Да, Нет, true, false, 1 и 0'),
            ('id', 'Оставьте пустым для нового питомца, укажите id для обновления карточки'),
        ],
    }


@admin_required
def admin_pet_excel(request):
    return render(request, 'shelter/admin_pet_excel.html', build_pet_excel_context())


@admin_required
def admin_pet_template_excel(request):
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
    except ImportError:
        messages.error(request, 'Для работы с Excel нужно установить openpyxl.')
        return redirect('admin_pet_excel')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Шаблон питомцев'

    worksheet.append(PET_EXCEL_HEADERS)
    worksheet.append([PET_EXCEL_EXAMPLE[header] for header in PET_EXCEL_HEADERS])

    for index, column in enumerate(PET_EXCEL_COLUMNS, start=1):
        header_cell = worksheet.cell(row=1, column=index)
        header_cell.comment = Comment(column['hint'], 'Лапки')

    help_sheet = workbook.create_sheet('Подсказки')
    help_sheet.append(['Колонка', 'Что указывать', 'Пример'])
    for column in PET_EXCEL_COLUMNS:
        help_sheet.append([column['key'], column['hint'], PET_EXCEL_EXAMPLE[column['key']]])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="pets_template.xlsx"'
    return response


@admin_required
def admin_pet_export_excel(request):
    try:
        from openpyxl import Workbook
    except ImportError:
        messages.error(request, 'Для работы с Excel нужно установить openpyxl.')
        return redirect('admin_pet_excel')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Питомцы'
    worksheet.append(PET_EXCEL_HEADERS)

    for pet in Pet.objects.order_by('id'):
        worksheet.append(
            [
                pet.id,
                pet.name,
                pet.get_type_display(),
                pet.get_gender_display(),
                pet.age_months,
                pet.get_age_group_display(),
                pet.description,
                pet_excel_bool_label(pet.sterilized),
                pet_excel_bool_label(pet.vaccinated),
                pet_excel_bool_label(pet.adopted),
                pet_excel_bool_label(pet.can_live_with_children),
                pet_excel_bool_label(pet.can_live_with_other_pets),
                pet.get_activity_level_display(),
                pet_excel_bool_label(pet.suitable_for_apartment),
                pet_excel_bool_label(pet.requires_experience),
                pet.get_temperament_display(),
                pet.get_health_status_display(),
                pet_excel_bool_label(pet.can_stay_alone),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="pets_export.xlsx"'
    return response


@admin_required
def admin_pet_import_excel(request):
    if request.method != 'POST':
        return redirect('admin_pet_excel')

    form = PetImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, 'Выберите Excel-файл для загрузки.')
        return render(request, 'shelter/admin_pet_excel.html', build_pet_excel_context(form))

    try:
        from openpyxl import load_workbook
    except ImportError:
        messages.error(request, 'Для работы с Excel нужно установить openpyxl.')
        return redirect('admin_pet_excel')

    workbook = load_workbook(form.cleaned_data['file'])
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        messages.error(request, 'Файл пустой.')
        return redirect('admin_pet_excel')

    headers = [str(value).strip() if value is not None else '' for value in rows[0]]
    index = {name: position for position, name in enumerate(headers)}
    required_headers = {'name', 'type', 'gender', 'age_months', 'age_group'}

    if not required_headers.issubset(index.keys()):
        messages.error(request, 'В файле не хватает обязательных колонок для импорта.')
        return redirect('admin_pet_excel')

    imported = 0
    updated = 0

    def value_from(row, key, default=''):
        position = index.get(key)
        if position is None or position >= len(row):
            return default
        value = row[position]
        return default if value is None else value

    def bool_from(row, key, default=False):
        value = value_from(row, key, default)
        return pet_excel_bool_from_value(value, default)

    for row in rows[1:]:
        if not any(value is not None and str(value).strip() for value in row):
            continue

        pet_id = value_from(row, 'id', None)
        defaults = {
            'name': str(value_from(row, 'name', '')).strip(),
            'type': pet_excel_normalize_choice(
                value_from(row, 'type', 'cat'),
                PET_EXCEL_CHOICE_MAPPINGS['type'],
                'cat',
            ),
            'gender': pet_excel_normalize_choice(
                value_from(row, 'gender', 'male'),
                PET_EXCEL_CHOICE_MAPPINGS['gender'],
                'male',
            ),
            'age_months': int(value_from(row, 'age_months', 0) or 0),
            'age_group': pet_excel_normalize_choice(
                value_from(row, 'age_group', 'adult'),
                PET_EXCEL_CHOICE_MAPPINGS['age_group'],
                'adult',
            ),
            'description': str(value_from(row, 'description', '')).strip(),
            'sterilized': bool_from(row, 'sterilized'),
            'vaccinated': bool_from(row, 'vaccinated'),
            'adopted': bool_from(row, 'adopted'),
            'can_live_with_children': bool_from(row, 'can_live_with_children', True),
            'can_live_with_other_pets': bool_from(row, 'can_live_with_other_pets', True),
            'activity_level': pet_excel_normalize_choice(
                value_from(row, 'activity_level', 'medium'),
                PET_EXCEL_CHOICE_MAPPINGS['activity_level'],
                'medium',
            ),
            'suitable_for_apartment': bool_from(row, 'suitable_for_apartment', True),
            'requires_experience': bool_from(row, 'requires_experience'),
            'temperament': pet_excel_normalize_choice(
                value_from(row, 'temperament', 'friendly'),
                PET_EXCEL_CHOICE_MAPPINGS['temperament'],
                'friendly',
            ),
            'health_status': pet_excel_normalize_choice(
                value_from(row, 'health_status', 'healthy'),
                PET_EXCEL_CHOICE_MAPPINGS['health_status'],
                'healthy',
            ),
            'can_stay_alone': bool_from(row, 'can_stay_alone', True),
        }

        if pet_id not in (None, '') and Pet.objects.filter(pk=int(pet_id)).exists():
            Pet.objects.filter(pk=int(pet_id)).update(**defaults)
            updated += 1
        else:
            Pet.objects.create(**defaults)
            imported += 1

    messages.success(request, f'Импорт завершён: добавлено {imported}, обновлено {updated}.')
    return redirect('admin_pet_list')


@admin_required
def admin_pet_list(request):
    pets = Pet.objects.order_by('adopted', 'name')
    return render(
        request,
        'shelter/admin_pet_list.html',
        {
            'pets': pets,
            'page_title': 'Управление питомцами',
        },
    )


@admin_required
def admin_pet_create(request):
    if request.method == 'POST':
        form = PetAdminForm(request.POST, request.FILES)
        if form.is_valid():
            pet = form.save()
            messages.success(request, f'Карточка питомца «{pet.name}» создана.')
            return redirect('admin_pet_list')
    else:
        form = PetAdminForm()

    return render(
        request,
        'shelter/admin_pet_form.html',
        {
            'form': form,
            'page_title': 'Добавить питомца',
            'submit_label': 'Сохранить питомца',
        },
    )


@admin_required
def admin_pet_edit(request, pk):
    pet = get_object_or_404(Pet, pk=pk)

    if request.method == 'POST':
        form = PetAdminForm(request.POST, request.FILES, instance=pet)
        if form.is_valid():
            pet = form.save()
            messages.success(request, f'Карточка питомца «{pet.name}» обновлена.')
            return redirect('admin_pet_list')
    else:
        form = PetAdminForm(instance=pet)

    return render(
        request,
        'shelter/admin_pet_form.html',
        {
            'form': form,
            'page_title': f'Редактирование: {pet.name}',
            'submit_label': 'Сохранить изменения',
            'pet': pet,
        },
    )


@admin_required
def admin_news_list(request):
    news_items = News.objects.order_by('-date')
    return render(
        request,
        'shelter/admin_news_list.html',
        {
            'news_items': news_items,
            'page_title': 'Управление новостями',
        },
    )


@admin_required
def admin_news_create(request):
    if request.method == 'POST':
        form = NewsAdminForm(request.POST, request.FILES)
        if form.is_valid():
            news_item = form.save()
            messages.success(request, f'Новость «{news_item.title}» создана.')
            return redirect('admin_news_list')
    else:
        form = NewsAdminForm()

    return render(
        request,
        'shelter/admin_news_form.html',
        {
            'form': form,
            'page_title': 'Добавить новость',
            'submit_label': 'Сохранить новость',
        },
    )


@admin_required
def admin_news_edit(request, pk):
    news_item = get_object_or_404(News, pk=pk)

    if request.method == 'POST':
        form = NewsAdminForm(request.POST, request.FILES, instance=news_item)
        if form.is_valid():
            news_item = form.save()
            messages.success(request, f'Новость «{news_item.title}» обновлена.')
            return redirect('admin_news_list')
    else:
        form = NewsAdminForm(instance=news_item)

    return render(
        request,
        'shelter/admin_news_form.html',
        {
            'form': form,
            'page_title': f'Редактирование новости: {news_item.title}',
            'submit_label': 'Сохранить изменения',
            'news_item': news_item,
        },
    )


@admin_required
def admin_application_list(request):
    applications = AdoptionApplication.objects.select_related('pet', 'user').order_by('-created_at')
    paginator = Paginator(applications, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(
        request,
        'shelter/admin_application_list.html',
        {
            'page_obj': page_obj,
            'page_title': 'Управление заявками',
        },
    )


@admin_required
def admin_application_edit(request, pk):
    application = get_object_or_404(AdoptionApplication.objects.select_related('pet', 'user'), pk=pk)

    if request.method == 'POST':
        form = AdoptionApplicationAdminForm(request.POST, instance=application)
        if form.is_valid():
            form.save()
            messages.success(request, f'Заявка для питомца «{application.pet.name}» обновлена.')
            return redirect('admin_application_list')
    else:
        form = AdoptionApplicationAdminForm(instance=application)

    return render(
        request,
        'shelter/admin_application_form.html',
        {
            'form': form,
            'application': application,
            'page_title': f'Заявка на {application.pet.name}',
            'submit_label': 'Сохранить заявку',
        },
    )


@admin_required
def admin_feedback_list(request):
    feedback_items = Feedback.objects.order_by('-submitted_at')
    return render(
        request,
        'shelter/admin_feedback_list.html',
        {
            'feedback_items': feedback_items,
            'page_title': 'Сообщения с сайта',
        },
    )


